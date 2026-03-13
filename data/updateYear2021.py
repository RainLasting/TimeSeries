from pathlib import Path

import pandas as pd

in_path = "anno_data9.0.xlsx"
out_path = "anno_data9.0_2021.xlsx"

# 读取 Excel
# 注意：如果你的环境没有 openpyxl，先在终端运行：pip install openpyxl

df = pd.read_excel(in_path)
df.head()

# 将 time 列年份统一改为 2021（保留月/日）

if "time" not in df.columns:
    raise KeyError(f"找不到 'time' 列，实际列名: {list(df.columns)}")

s = df["time"]

# 兼容两类情况：
# 1) Excel 里是日期/时间类型（pandas 会读成 datetime64）
# 2) Excel 里是字符串（例如 '3/14/25'）

if pd.api.types.is_datetime64_any_dtype(s):
    dt = s
else:
    # 保留原值用于回填
    dt = pd.to_datetime(s.astype(str).str.strip(), errors="coerce", format="%m/%d/%y")
    # 如果存在非 m/d/yy 的格式，尝试再做一次更宽松的解析
    mask = dt.isna() & s.notna()
    if mask.any():
        dt2 = pd.to_datetime(s.astype(str).str.strip(), errors="coerce")
        dt = dt.fillna(dt2)

# 只对可解析为日期的行改年份；并且最终只保留 YYYY-MM-DD（不带小时）

# 统一为 pandas datetime，并截断到“日期”（去掉时分秒）
dt = pd.to_datetime(dt, errors="coerce")
dt_date = dt.dt.normalize()

mask_ok = dt_date.notna()
updated = s.copy()

# 替换年份（值保留为日期/时间类型，写出时再用 Excel 格式隐藏时间）
updated.loc[mask_ok] = dt_date.loc[mask_ok].apply(lambda x: x.replace(year=2021))

df_out = df.copy()
df_out["time"] = updated

# 前后对比（前 10 行）
preview = pd.DataFrame({"before": s.head(10), "after": df_out["time"].head(10)})
preview

# 写出到新的 Excel，并强制 Excel 显示为“年-月-日”

# 先写出
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Sheet1")

# 再用 openpyxl 设置 time 列的单元格格式，避免 Excel 显示 00:00:00
from openpyxl import load_workbook

wb = load_workbook(out_path)
ws = wb["Sheet1"]

# 找到 time 列位置
header = [cell.value for cell in ws[1]]
if "time" not in header:
    raise KeyError(f"导出后找不到 'time' 列，表头为: {header}")

time_col_idx = header.index("time") + 1  # 1-based

for r in range(2, ws.max_row + 1):
    c = ws.cell(row=r, column=time_col_idx)
    # 只要是日期/时间类型，就统一设置显示格式
    c.number_format = "yyyy-mm-dd"

wb.save(out_path)
print(f"Saved: {out_path.resolve()}")